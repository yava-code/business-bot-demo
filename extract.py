import fitz  # PyMuPDF
import openpyxl


def extract_pdf(path: str) -> str:
    doc = fitz.open(path)
    pages = [page.get_text() for page in doc]
    doc.close()
    return "\n".join(pages).strip()


def extract_excel(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"=== Sheet: {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                lines.append("\t".join("" if v is None else str(v) for v in row))
    return "\n".join(lines).strip()
